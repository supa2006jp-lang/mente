import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed.")
    sys.exit(1)

try:
    file_path = r"C:\Users\PC_User\Desktop\g\てんき.xlsx"
    out_path = r"C:\Users\PC_User\.gemini\antigravity\brain\80205a77-87ca-4c37-be1c-0be5eda2535c\results.txt"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"Sheets: {wb.sheetnames}\n")
        ws = wb.active
        f.write(f"Active: {ws.title}\n")
        
        for row in ws.iter_rows(max_row=50, max_col=20):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).replace('\n', ' ')
                    f.write(f"[{cell.coordinate}] {val}\n")
    print("Success")
except Exception as e:
    print(f"Error: {e}")
sys.exit(0)
